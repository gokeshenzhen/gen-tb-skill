#!/usr/bin/env python3
"""
gen-tb eval harness — runs scaffold.py against each fixture and checks
the assertions in evals/evals.json. Writes a benchmark.json with per-eval
pass/fail + timing.

Usage:
    python3 scripts/run_evals.py [--with-skill | --baseline]
                                  [--out evals/iteration-N/]
                                  [--filter <eval_name>]

This harness mirrors skill-creator's eval framework but is lightweight —
no subagent spawn, no LLM grading. For gen-tb the work is mechanical:

  with_skill:  cd to a fresh workspace, copy the fixture's spec/RTL,
               write a hand-crafted intake.yaml (since intake interview
               is done by the model not by this harness), run scaffold.py,
               compile, run sanity, check assertions.

  baseline:    skip scaffold.py; emit nothing. Compile & sanity fail
               (no makefile generated). This baseline isn't very useful
               for v1.1; it'll be more interesting once we have a
               with-skill-but-vanilla-Claude comparison.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import os
import re
import shutil
import subprocess
import sys
import time
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
FIXTURES = ROOT / "evals" / "fixtures"


def _fixture_hash(fixture_dir: Path) -> str:
    """Hash all regular files in the fixture to detect pollution."""
    h = hashlib.sha256()
    for p in sorted(fixture_dir.rglob("*")):
        if p.is_file():
            h.update(str(p.relative_to(fixture_dir)).encode())
            h.update(p.read_bytes())
    return h.hexdigest()


def _prepare_workspace(eval_def: dict, scratch_root: Path) -> Path:
    fixture = FIXTURES / eval_def["fixture"]
    ip_root = scratch_root / eval_def["fixture"]
    if ip_root.exists():
        shutil.rmtree(ip_root)
    ip_root.mkdir(parents=True)
    # Symlink the read-only inputs from the fixture
    (ip_root / "spec").symlink_to(fixture / "spec")
    (ip_root / "rtl").symlink_to(fixture / "rtl")
    if (fixture / "ref_model").exists():
        (ip_root / "ref_model").symlink_to(fixture / "ref_model")
    (ip_root / ".prj_top").touch()
    return ip_root


def _emit_audit_inputs(eval_def: dict, ip_root: Path):
    """Hand-craft the audit inputs that the model would normally produce
    during Phase 1-3. The harness uses the canonical inputs captured per
    fixture under evals/fixtures/<name>/expected/."""
    audit = ip_root / "work" / "_gen_audit"
    norm = audit / "spec_normalized"
    norm.mkdir(parents=True, exist_ok=True)

    name = eval_def["fixture"]
    fixture_root = FIXTURES / name
    expected = fixture_root / "expected"

    # intake.yaml — re-root any references to the source ip_root
    intake_content = (expected / "intake.yaml").read_text()
    # rewrite ip_root references so paths point to this run's ip_root
    intake_content = re.sub(r"ip_root:.*", f"ip_root: {ip_root}", intake_content)
    intake_content = re.sub(
        r"\$PROJ_DIR/(ref_model_src|ref_model)/",
        "$PROJ_DIR/ref_model/",
        intake_content,
    )
    (audit / "intake.yaml").write_text(intake_content)

    # rtl_discovery.yaml — re-emit with current paths
    _emit_rtl_discovery(eval_def, ip_root, audit)

    # registers.yaml — parse from the fixture xlsx
    xlsx_candidates = list((fixture_root / "spec").glob("*_regs.xlsx"))
    if xlsx_candidates:
        _parse_xlsx_to_yaml(xlsx_candidates[0], norm / "registers.yaml")


def _emit_rtl_discovery(eval_def: dict, ip_root: Path, audit: Path):
    """Discover RTL files; emit minimal rtl_discovery.yaml."""
    name = eval_def["fixture"]
    rtl_files = sorted((ip_root / "rtl").glob("*.v"))
    # Find top: module not instantiated by another
    instantiated = set()
    for f in rtl_files:
        txt = f.read_text()
        for other in rtl_files:
            mod = other.stem
            if mod != f.stem and re.search(rf"\b{mod}\s+\w+\s*\(", txt):
                instantiated.add(mod)
    candidates = [f.stem for f in rtl_files if f.stem not in instantiated]
    # prefer name containing "apb_wrap"
    top = next((c for c in candidates if "apb_wrap" in c), candidates[0] if candidates else "")

    # crude topo order: leaves first, top last
    order = []
    seen = set()
    for f in rtl_files:
        if f.stem == top:
            continue
        order.append(f)
        seen.add(f.stem)
    if top:
        order.append(ip_root / "rtl" / f"{top}.v")

    lines = [
        "mode: scan",
        f"ip_name: {name}",
        f"ip_root: {ip_root}",
        "rtl_dir: $PROJ_DIR/rtl",
        "filelist_origin: generated",
        "top_module:",
        f"  name: {top}",
        f"  file: $PROJ_DIR/rtl/{top}.v",
        "  confidence: medium",
        "files:",
    ]
    for i, f in enumerate(order, 1):
        lines.append(
            f"  - {{path: $PROJ_DIR/rtl/{f.name}, role: {'top' if f.stem == top else 'leaf'}, order: {i}}}"
        )
    lines += [
        "apb_interface:",
        "  pclk: pclk",
        "  presetn: presetn",
        "  psel: psel",
        "  penable: penable",
        "  pwrite: pwrite",
        "  paddr:  {name: paddr,  width: 12}",
        "  pwdata: {name: pwdata, width: 32}",
        "  prdata: {name: prdata, width: 32}",
        "  pready: pready",
        "  pslverr: pslverr",
        "other_pads: []" if name == "aes128" else "other_pads:",
    ]
    if name != "aes128":
        lines += [
            "  - {name: irq,       dir: out, role: interrupt}",
            "  - {name: stx_pad_o, dir: out, role: serial_tx}",
            "  - {name: srx_pad_i, dir: in,  role: serial_rx}",
            "  - {name: rts_pad_o, dir: out, role: flow_ctrl}",
            "  - {name: cts_pad_i, dir: in,  role: flow_ctrl}",
            "  - {name: dtr_pad_o, dir: out, role: modem}",
            "  - {name: dsr_pad_i, dir: in,  role: modem}",
            "  - {name: ri_pad_i,  dir: in,  role: modem}",
            "  - {name: dcd_pad_i, dir: in,  role: modem}",
        ]
    (audit / "rtl_discovery.yaml").write_text("\n".join(lines) + "\n")


def _parse_xlsx_to_yaml(xlsx: Path, out_yaml: Path):
    """Minimal inline xlsx → registers.yaml normalizer."""
    from openpyxl import load_workbook
    wb = load_workbook(xlsx, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    regs = {}
    for r in rows:
        if r[0] is None:
            continue
        name, off, width, access, reset, fn, fb, fa, fr, desc = r
        key = (name, off)
        if key not in regs:
            regs[key] = {
                "name": name, "offset": int(off), "width": int(width),
                "access": access, "reset": int(reset), "fields": [],
            }
        regs[key]["fields"].append({
            "name": fn, "bits": fb, "access": fa,
            "reset": int(fr) if fr is not None else 0,
            "desc": desc or "",
        })
    out = ["registers:"]
    for r in regs.values():
        out.append(f"  - name: {r['name']}")
        out.append(f"    offset: 0x{r['offset']:02X}")
        out.append(f"    width: {r['width']}")
        out.append(f"    access: {r['access']}")
        out.append(f"    reset: 0x{r['reset']:X}")
        out.append( "    fields:")
        for f in r["fields"]:
            out.append(f"      - name: {f['name']}")
            out.append(f"        bits: \"{f['bits']}\"")
            out.append(f"        access: {f['access']}")
            out.append(f"        reset: 0x{f['reset']:X}")
    out_yaml.write_text("\n".join(out) + "\n")


def _run_scaffold(ip_root: Path) -> tuple[bool, str]:
    result = subprocess.run(
        [sys.executable, str(ROOT / "scripts" / "scaffold.py"),
         "--ip-root", str(ip_root), "--force"],
        capture_output=True, text=True,
    )
    return result.returncode == 0, result.stdout + result.stderr


def _run_make(ip_root: Path, target: str, sv_case: str | None = None) -> dict:
    env = os.environ.copy()
    env["PROJ_DIR"] = str(ip_root)
    env["WORK_DIR"] = str(ip_root / "work")
    if env.get("VCS_HOME") and not env.get("UVM_HOME"):
        env["UVM_HOME"] = env["VCS_HOME"] + "/etc/uvm-1.2"

    cmd = ["make", "-f", str(ip_root / "script" / "makefile"), target]
    if sv_case:
        cmd.append(f"SV_CASE={sv_case}")
    t0 = time.time()
    result = subprocess.run(cmd, capture_output=True, text=True, env=env, cwd=ip_root)
    return {
        "rc": result.returncode,
        "stdout": result.stdout,
        "stderr": result.stderr,
        "duration_s": round(time.time() - t0, 2),
    }


def _read_log(ip_root: Path, sv_case: str, kind: str) -> str:
    fname = "run.log" if kind == "sim" else "comp.log"
    p = ip_root / "work" / f"work_{sv_case}_" / fname
    return p.read_text() if p.exists() else ""


def _check_assertion(a: dict, ctx: dict) -> tuple[bool, str]:
    """Returns (passed, evidence)."""
    kind = a["kind"]
    ip_root = ctx["ip_root"]
    eval_def = ctx["eval_def"]

    if kind == "files_exist":
        files_txt = ROOT / a["files_txt"]
        missing = []
        for line in files_txt.read_text().splitlines():
            line = line.strip()
            if not line or line.startswith("#"):
                continue
            parts = line.split("\t") if "\t" in line else line.split(None, 1)
            path = parts[0]
            kind_label = parts[1] if len(parts) > 1 else "nonempty"
            target = ip_root / path
            if not target.exists():
                missing.append(f"{path} ({kind_label})")
        if missing:
            return False, f"missing: {', '.join(missing[:5])}{'...' if len(missing)>5 else ''}"
        return True, "all expected files present"

    if kind == "compile_exit_zero":
        rc = ctx["compile"]["rc"]
        return rc == 0, f"make comp rc={rc}"

    if kind == "compile_no_warnings":
        log = _read_log(ip_root, ctx["sanity_test"], "comp")
        ignore = a.get("ignore_patterns", [])
        warns = []
        for line in log.splitlines():
            if "warning" in line.lower() or "Warning" in line:
                if any(re.search(p, line) for p in ignore):
                    continue
                warns.append(line.strip())
        if warns:
            return False, f"{len(warns)} warning(s); first: {warns[0][:120]}"
        return True, "0 warnings"

    if kind == "compile_log_contains":
        log = _read_log(ip_root, ctx["sanity_test"], "comp")
        # also look at compile stdout in case log isn't there
        text = log + ctx["compile"]["stdout"] + ctx["compile"]["stderr"]
        return (a["needle"] in text), f"needle={a['needle']!r}"

    if kind == "sim_passes":
        run = ctx["sim_runs"].get(a["test"], {})
        if not run:
            return False, "test not run"
        if run["rc"] != 0:
            return False, f"make rc={run['rc']}"
        log = _read_log(ip_root, a["test"], "sim")
        m_err = re.search(r"UVM_ERROR\s*:\s*(\d+)", log)
        m_fat = re.search(r"UVM_FATAL\s*:\s*(\d+)", log)
        if m_err and int(m_err.group(1)) > 0:
            return False, f"UVM_ERROR={m_err.group(1)}"
        if m_fat and int(m_fat.group(1)) > 0:
            return False, f"UVM_FATAL={m_fat.group(1)}"
        return True, "UVM_ERROR=0 UVM_FATAL=0"

    if kind == "log_contains":
        log = _read_log(ip_root, a["test"], "sim")
        return (a["needle"] in log), f"needle={a['needle']!r}"

    if kind == "fixture_unchanged":
        before = ctx["fixture_hash_before"]
        after = _fixture_hash(FIXTURES / eval_def["fixture"])
        return before == after, "hash match" if before == after else "fixture modified!"

    return False, f"unknown assertion kind: {kind}"


def run_one(eval_def: dict, scratch_root: Path, mode: str) -> dict:
    t0 = time.time()
    fixture_dir = FIXTURES / eval_def["fixture"]
    ctx = {
        "eval_def": eval_def,
        "fixture_hash_before": _fixture_hash(fixture_dir),
        "sim_runs": {},
    }

    # 1. Set up workspace
    ip_root = _prepare_workspace(eval_def, scratch_root)
    ctx["ip_root"] = ip_root

    if mode == "with-skill":
        _emit_audit_inputs(eval_def, ip_root)
        scaffold_ok, scaffold_log = _run_scaffold(ip_root)
        if not scaffold_ok:
            return {
                "id": eval_def["id"], "name": eval_def["name"], "mode": mode,
                "passed": False, "stage": "scaffold", "log": scaffold_log[-1000:],
                "duration_s": round(time.time() - t0, 2),
            }

    # 2. Identify the sanity test name (best-effort)
    ctx["sanity_test"] = f"{eval_def['fixture']}_sanity_test"

    # 3. Compile
    ctx["compile"] = _run_make(ip_root, "comp", ctx["sanity_test"])
    # 4. Run each test referenced by assertions
    tests_to_run = sorted({a["test"] for a in eval_def["assertions"] if "test" in a})
    for t in tests_to_run:
        ctx["sim_runs"][t] = _run_make(ip_root, "all", t)

    # 5. Check assertions
    results = []
    for a in eval_def["assertions"]:
        passed, evidence = _check_assertion(a, ctx)
        results.append({"text": a["name"], "passed": passed, "evidence": evidence})

    overall = all(r["passed"] for r in results)
    return {
        "id": eval_def["id"], "name": eval_def["name"], "mode": mode,
        "passed": overall,
        "expectations": results,
        "compile_rc": ctx["compile"]["rc"],
        "compile_duration_s": ctx["compile"]["duration_s"],
        "sim_durations_s": {t: r["duration_s"] for t, r in ctx["sim_runs"].items()},
        "duration_s": round(time.time() - t0, 2),
    }


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--mode", choices=["with-skill", "baseline"], default="with-skill")
    ap.add_argument("--out", type=Path, default=ROOT / "evals" / "iteration-1")
    ap.add_argument("--filter", help="run only this eval name")
    ap.add_argument("--scratch", type=Path, default=Path("/tmp/gen-tb-evals"))
    args = ap.parse_args()

    if args.scratch.exists():
        shutil.rmtree(args.scratch)
    args.scratch.mkdir(parents=True)
    args.out.mkdir(parents=True, exist_ok=True)

    spec = json.loads((ROOT / "evals" / "evals.json").read_text())
    runs = []
    for eval_def in spec["evals"]:
        if args.filter and eval_def["name"] != args.filter:
            continue
        print(f"  [run] {eval_def['name']} ({args.mode})")
        r = run_one(eval_def, args.scratch, args.mode)
        print(f"     → {'PASS' if r['passed'] else 'FAIL'} ({r['duration_s']}s)")
        for e in r.get("expectations", []):
            mark = "✓" if e["passed"] else "✗"
            print(f"        {mark} {e['text']:35s} {e['evidence']}")
        runs.append(r)

    bench = {
        "skill": spec["skill_name"],
        "mode": args.mode,
        "n_evals": len(runs),
        "n_passed": sum(1 for r in runs if r["passed"]),
        "runs": runs,
    }
    out = args.out / f"benchmark-{args.mode}.json"
    out.write_text(json.dumps(bench, indent=2))
    print(f"\nbenchmark saved: {out}")
    print(f"summary: {bench['n_passed']}/{bench['n_evals']} passed")
    return 0 if bench["n_passed"] == bench["n_evals"] else 1


if __name__ == "__main__":
    sys.exit(main())
