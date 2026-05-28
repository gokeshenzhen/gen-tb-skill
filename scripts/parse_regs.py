#!/usr/bin/env python3
"""Phase 3 helper: normalize register tables into registers.yaml.

Supported inputs:
- xlsx/csv tables using normalized headers or the fixture column order
- markdown pipe tables
- a practical subset of IP-XACT XML register maps
"""

from __future__ import annotations

import argparse
import csv
import re
import xml.etree.ElementTree as ET
from collections import OrderedDict
from pathlib import Path
from typing import Any

import yaml


FIXTURE_COLUMNS = [
    "name", "offset", "width", "access", "reset",
    "field_name", "field_bits", "field_access", "field_reset", "desc",
]

HEADER_ALIASES = {
    "name": {"name", "reg", "register", "registername", "regname"},
    "offset": {"offset", "address", "addr", "baseaddress", "addressoffset"},
    "width": {"width", "regwidth", "size", "bits"},
    "access": {"access", "type", "rw", "permission"},
    "reset": {"reset", "resetvalue", "resetval", "rst", "default"},
    "field_name": {"field", "fieldname", "bitname", "field_name", "nameoffield"},
    "field_bits": {"fieldbits", "field_bits", "bits", "bit", "bitrange", "range"},
    "field_access": {"fieldaccess", "field_access", "fieldrw", "fieldpermission"},
    "field_reset": {"fieldreset", "field_reset", "fielddefault"},
    "desc": {"desc", "description", "fielddesc", "fielddescription"},
    "aliased_by": {"aliasedby", "alias", "bankselect", "bank"},
    "aliased_by_value": {"aliasedbyvalue", "aliasvalue", "bankvalue"},
    "array_of": {"arrayof", "array", "count", "dim"},
    "stride": {"stride", "arraystride"},
    "effect": {"effect", "sideeffect", "sideeffects"},
}


class ParseReport:
    def __init__(self) -> None:
        self.inputs: list[str] = []
        self.selected_sources: list[str] = []
        self.assumptions: list[str] = []
        self.conflicts: list[str] = []
        self.reset_mismatches: list[str] = []
        self.alias_decisions: list[str] = []
        self.array_folding: list[str] = []
        self.side_effects: list[str] = []
        self.access_normalization: list[str] = []
        self.warnings: list[str] = []

    def write(self, path: Path) -> None:
        sections = [
            ("Inputs", self.inputs),
            ("Selected Sources", self.selected_sources),
            ("Assumptions", self.assumptions),
            ("Conflicts", self.conflicts),
            ("Reset value mismatches", self.reset_mismatches),
            ("Alias decisions", self.alias_decisions),
            ("Array folding", self.array_folding),
            ("Side-effect inference", self.side_effects),
            ("Access normalization", self.access_normalization),
            ("Reference model decision", [
                "Register-level expected behavior belongs in the generated RAL/reg block."
            ]),
            ("Warnings", self.warnings),
        ]
        lines = ["# Parse Report", ""]
        for title, items in sections:
            lines += [f"## {title}"]
            lines += [f"- {item}" for item in items] if items else ["- none"]
            lines.append("")
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_text("\n".join(lines).rstrip() + "\n")


def _norm_header(raw: Any) -> str:
    return re.sub(r"[^a-z0-9]", "", str(raw or "").lower())


def _canonical_header(raw: Any) -> str | None:
    norm = _norm_header(raw)
    for canon, aliases in HEADER_ALIASES.items():
        if norm in aliases:
            return canon
    return None


def _parse_int(value: Any, default: int | None = None) -> int:
    if value is None or value == "":
        if default is not None:
            return default
        raise ValueError("missing integer value")
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        if value.is_integer():
            return int(value)
        raise ValueError(f"non-integer numeric value: {value!r}")
    text = str(value).strip().replace("_", "")
    text = text.strip("[]")
    if not text:
        if default is not None:
            return default
        raise ValueError("missing integer value")
    m = re.match(r"^(?:\d+)?'([hdb])([0-9a-fA-FxXzZ]+)$", text)
    if m:
        base = {"h": 16, "d": 10, "b": 2}[m.group(1).lower()]
        digits = re.sub(r"[xXzZ]", "0", m.group(2))
        return int(digits, base)
    if text.lower().startswith("0x"):
        return int(text, 16)
    if re.search(r"[a-fA-F]", text):
        return int(text, 16)
    return int(text, 10)


UVM12_FIELD_ACCESSES = {
    "RO", "RW", "RC", "RS", "WRC", "WRS", "WC", "WS", "WSRC", "WCRS",
    "W1C", "W1S", "W1T", "W0C", "W0S", "W0T",
    "W1SRC", "W1CRS", "W0SRC", "W0CRS",
    "WO", "WOC", "WOS", "W1", "WO1",
}

WRITE_CAPABLE_FIELD_ACCESSES = {
    "RW", "WRC", "WRS", "WC", "WS", "WSRC", "WCRS",
    "W1C", "W1S", "W1T", "W0C", "W0S", "W0T",
    "W1SRC", "W1CRS", "W0SRC", "W0CRS",
    "WO", "WOC", "WOS", "W1", "WO1",
}

READ_CAPABLE_FIELD_ACCESSES = {
    "RO", "RW", "RC", "RS", "WRC", "WRS", "WC", "WS", "WSRC", "WCRS",
    "W1C", "W1S", "W1T", "W0C", "W0S", "W0T",
    "W1SRC", "W1CRS", "W0SRC", "W0CRS", "W1",
}

ACCESS_ALIASES = {
    "R": "RO",
    "RD": "RO",
    "RO": "RO",
    "READ": "RO",
    "READONLY": "RO",
    "W": "WO",
    "WO": "WO",
    "WRITE": "WO",
    "WRITEONLY": "WO",
    "RW": "RW",
    "WR": "RW",
    "READWRITE": "RW",
    "WRITEREAD": "RW",
    "READWRITABLE": "RW",
    "RC": "RC",
    "READCLEAR": "RC",
    "READCLEARS": "RC",
    "READTOCLEAR": "RC",
    "RS": "RS",
    "READSET": "RS",
    "READSETS": "RS",
    "READTOSET": "RS",
    "WC": "WC",
    "WRITECLEAR": "WC",
    "WRITECLEARS": "WC",
    "WRITETOCLEAR": "WC",
    "WS": "WS",
    "WRITESET": "WS",
    "WRITESETS": "WS",
    "WRITETOSET": "WS",
    "WRC": "WRC",
    "WRITEREADCLEAR": "WRC",
    "WRITEREADTOCLEAR": "WRC",
    "WRS": "WRS",
    "WRITEREADSET": "WRS",
    "WRITEREADTOSET": "WRS",
    "WSRC": "WSRC",
    "WRITESETREADCLEAR": "WSRC",
    "WRITESETREADTOCLEAR": "WSRC",
    "WCRS": "WCRS",
    "WRITECLEARREADSET": "WCRS",
    "WRITECLEARREADTOSET": "WCRS",
    "W1C": "W1C",
    "WRITE1C": "W1C",
    "WRITE1CLEAR": "W1C",
    "WRITE1TOCLEAR": "W1C",
    "WRITEONETOCLEAR": "W1C",
    "WRITEONECLEAR": "W1C",
    "W1S": "W1S",
    "WRITE1S": "W1S",
    "WRITE1SET": "W1S",
    "WRITE1TOSET": "W1S",
    "WRITEONETOSET": "W1S",
    "WRITEONESET": "W1S",
    "W1T": "W1T",
    "WRITE1T": "W1T",
    "WRITE1TOGGLE": "W1T",
    "WRITE1TOTOGGLE": "W1T",
    "WRITEONETOTOGGLE": "W1T",
    "WRITEONETOGGLE": "W1T",
    "W0C": "W0C",
    "WRITE0C": "W0C",
    "WRITE0CLEAR": "W0C",
    "WRITE0TOCLEAR": "W0C",
    "WRITEZEROTOCLEAR": "W0C",
    "WRITEZEROCLEAR": "W0C",
    "W0S": "W0S",
    "WRITE0S": "W0S",
    "WRITE0SET": "W0S",
    "WRITE0TOSET": "W0S",
    "WRITEZEROTOSET": "W0S",
    "WRITEZEROSET": "W0S",
    "W0T": "W0T",
    "WRITE0T": "W0T",
    "WRITE0TOGGLE": "W0T",
    "WRITE0TOTOGGLE": "W0T",
    "WRITEZEROTOTOGGLE": "W0T",
    "WRITEZEROTOGGLE": "W0T",
    "W1SRC": "W1SRC",
    "WRITE1SETREADCLEAR": "W1SRC",
    "WRITE1SETREADTOCLEAR": "W1SRC",
    "W1CRS": "W1CRS",
    "WRITE1CLEARREADSET": "W1CRS",
    "WRITE1CLEARREADTOSET": "W1CRS",
    "W0SRC": "W0SRC",
    "WRITE0SETREADCLEAR": "W0SRC",
    "WRITE0SETREADTOCLEAR": "W0SRC",
    "W0CRS": "W0CRS",
    "WRITE0CLEARREADSET": "W0CRS",
    "WRITE0CLEARREADTOSET": "W0CRS",
    "WOC": "WOC",
    "WRITEONLYCLEAR": "WOC",
    "WRITEONLYTOCLEAR": "WOC",
    "WOS": "WOS",
    "WRITEONLYSET": "WOS",
    "WRITEONLYTOSET": "WOS",
    "W1": "W1",
    "WRITEONCE": "W1",
    "WO1": "WO1",
    "WRITEONLYONCE": "WO1",
}


def _access_key(value: Any) -> str:
    return re.sub(r"[^A-Z0-9]", "", str(value or "").upper())


def _normalize_field_access(
    value: Any,
    default: str = "RW",
    report: ParseReport | None = None,
    context: str = "access",
) -> str:
    raw = str(value or default).strip()
    key = _access_key(raw)
    default_key = _access_key(default)
    canonical = ACCESS_ALIASES.get(key) or (
        key if key in UVM12_FIELD_ACCESSES else None
    )
    if canonical is None:
        fallback = ACCESS_ALIASES.get(default_key, "RW")
        if report is not None:
            report.warnings.append(
                f"{context}: unsupported access {raw!r}; using {fallback}"
            )
        return fallback
    if report is not None and key != canonical:
        report.access_normalization.append(f"{context}: {raw!r} -> {canonical}")
    return canonical


def _access_to_map_right(access: str) -> str:
    readable = access in READ_CAPABLE_FIELD_ACCESSES
    writable = access in WRITE_CAPABLE_FIELD_ACCESSES
    if readable and writable:
        return "RW"
    if readable:
        return "RO"
    if writable:
        return "WO"
    return "RO"


def _fields_to_map_right(fields: list[dict]) -> str:
    readable = any(f["access"] in READ_CAPABLE_FIELD_ACCESSES for f in fields)
    writable = any(f["access"] in WRITE_CAPABLE_FIELD_ACCESSES for f in fields)
    if readable and writable:
        return "RW"
    if readable:
        return "RO"
    if writable:
        return "WO"
    return "RO"


def _normalize_bits(value: Any, width: int = 32) -> str:
    if value is None or value == "":
        return f"{width - 1}:0"
    text = str(value).strip().replace("[", "").replace("]", "")
    text = text.replace(" ", "")
    if ":" in text:
        hi, lo = [int(x, 0) for x in text.split(":", 1)]
        return f"{max(hi, lo)}:{min(hi, lo)}"
    return str(int(text, 0))


def _bit_lsb_width(bits: str) -> tuple[int, int]:
    if ":" in bits:
        hi, lo = [int(x, 0) for x in bits.split(":", 1)]
        return min(hi, lo), abs(hi - lo) + 1
    bit = int(bits, 0)
    return bit, 1


def _sv_id(raw: Any, fallback: str) -> str:
    text = re.sub(r"\W", "_", str(raw or fallback).strip())
    if not text:
        text = fallback
    if re.match(r"\d", text):
        text = f"_{text}"
    return text


def _rows_from_xlsx(path: Path, report: ParseReport) -> list[dict[str, Any]]:
    from openpyxl import load_workbook

    wb = load_workbook(path, data_only=True)
    ws = wb.active
    raw_rows = list(ws.iter_rows(values_only=True))
    if not raw_rows:
        return []
    return _rows_from_matrix(raw_rows, f"{path.name}:{ws.title}", report)


def _rows_from_csv(path: Path, report: ParseReport) -> list[dict[str, Any]]:
    with path.open(newline="") as f:
        rows = list(csv.reader(f))
    return _rows_from_matrix(rows, path.name, report)


def _rows_from_markdown(path: Path, report: ParseReport) -> list[dict[str, Any]]:
    tables: list[list[list[str]]] = []
    current: list[list[str]] = []
    for line in path.read_text().splitlines():
        if "|" in line and re.match(r"^\s*\|?.*\|.*\|?\s*$", line):
            cells = [c.strip() for c in line.strip().strip("|").split("|")]
            if all(re.fullmatch(r":?-{3,}:?", c or "") for c in cells):
                continue
            current.append(cells)
        elif current:
            tables.append(current)
            current = []
    if current:
        tables.append(current)

    for table in tables:
        if not table:
            continue
        headers = [_canonical_header(c) for c in table[0]]
        if "name" in headers and "offset" in headers:
            return _rows_from_matrix(table, path.name, report)
    report.warnings.append(f"{path.name}: no markdown register table with name+offset headers found")
    return []


def _rows_from_matrix(matrix: list[Any], source: str, report: ParseReport) -> list[dict[str, Any]]:
    if not matrix:
        return []
    header = list(matrix[0])
    mapped = [_canonical_header(h) for h in header]
    if mapped.count("width") > 1:
        seen_width = False
        for i, col in enumerate(mapped):
            if col != "width":
                continue
            if seen_width:
                mapped[i] = "field_bits"
            else:
                seen_width = True
    has_header = "name" in mapped and "offset" in mapped
    if not has_header:
        if len(header) >= len(FIXTURE_COLUMNS):
            mapped = FIXTURE_COLUMNS + [None] * (len(header) - len(FIXTURE_COLUMNS))
            body = matrix[1:] if all(str(c or "").strip() for c in header[:3]) else matrix
            report.assumptions.append(f"{source}: used fixture column order")
        else:
            raise ValueError(f"{source}: no recognizable header row")
    else:
        body = matrix[1:]

    rows: list[dict[str, Any]] = []
    carry: dict[str, Any] = {}
    for idx, raw in enumerate(body, 2):
        if not any(c not in (None, "") for c in raw):
            continue
        row: dict[str, Any] = {}
        for col, val in zip(mapped, raw):
            if col:
                row[col] = val
        for key in ("name", "offset", "width", "access", "reset"):
            if row.get(key) in (None, "") and key in carry:
                row[key] = carry[key]
            elif row.get(key) not in (None, ""):
                carry[key] = row[key]
        row["_source"] = f"{source}:row{idx}"
        rows.append(row)
    return rows


def _local_name(tag: str) -> str:
    return tag.rsplit("}", 1)[-1]


def _children(elem: ET.Element, name: str) -> list[ET.Element]:
    return [c for c in list(elem) if _local_name(c.tag) == name]


def _child_text(elem: ET.Element, name: str, default: str | None = None) -> str | None:
    child = next((c for c in list(elem) if _local_name(c.tag) == name), None)
    if child is None or child.text is None:
        return default
    return child.text.strip()


def _rows_from_ipxact(path: Path, report: ParseReport) -> list[dict[str, Any]]:
    root = ET.parse(path).getroot()
    rows: list[dict[str, Any]] = []
    for reg in [e for e in root.iter() if _local_name(e.tag) == "register"]:
        reg_name = _child_text(reg, "name")
        offset = _child_text(reg, "addressOffset")
        if not reg_name or offset is None:
            continue
        width = _child_text(reg, "size", "32")
        access = _child_text(reg, "access", "read-write")
        reset = 0
        resets = next((c for c in reg.iter() if _local_name(c.tag) == "reset"), None)
        if resets is not None:
            reset = _parse_int(_child_text(resets, "value", "0"), 0)
        fields = _children(reg, "fields")
        field_nodes = []
        for fields_node in fields:
            field_nodes.extend(_children(fields_node, "field"))
        if not field_nodes:
            rows.append({
                "name": reg_name,
                "offset": offset,
                "width": width,
                "access": access,
                "reset": reset,
                "field_name": "data",
                "field_bits": f"{_parse_int(width) - 1}:0",
                "field_access": access,
                "field_reset": reset,
                "_source": f"{path.name}:{reg_name}",
            })
            continue
        for field in field_nodes:
            fname = _child_text(field, "name", "field")
            bit_offset = _parse_int(_child_text(field, "bitOffset", "0"), 0)
            bit_width = _parse_int(_child_text(field, "bitWidth", "1"), 1)
            faccess = _child_text(field, "access", access)
            # Field-level reset: use the field's own <reset><value> when the
            # IP-XACT declares one; otherwise leave it unset so the register
            # row resolves it by slicing the register-level reset (G8). A
            # hardcoded 0 here would masquerade as an explicit field reset.
            freset_node = next(
                (c for c in field.iter() if _local_name(c.tag) == "reset"), None
            )
            row = {
                "name": reg_name,
                "offset": offset,
                "width": width,
                "access": access,
                "reset": reset,
                "field_name": fname,
                "field_bits": f"{bit_offset + bit_width - 1}:{bit_offset}",
                "field_access": faccess,
                "desc": _child_text(field, "description", ""),
                "_source": f"{path.name}:{reg_name}.{fname}",
            }
            if freset_node is not None:
                row["field_reset"] = _child_text(freset_node, "value", "0")
            rows.append(row)
    report.selected_sources.append(f"{path.name}: parsed IP-XACT XML")
    return rows


def _rows_for_input(path: Path, report: ParseReport) -> list[dict[str, Any]]:
    suffix = path.suffix.lower()
    report.inputs.append(str(path))
    if suffix == ".xlsx":
        report.selected_sources.append(f"{path.name}: parsed xlsx")
        return _rows_from_xlsx(path, report)
    if suffix == ".csv":
        report.selected_sources.append(f"{path.name}: parsed csv")
        return _rows_from_csv(path, report)
    if suffix in {".md", ".markdown"}:
        report.selected_sources.append(f"{path.name}: parsed markdown")
        return _rows_from_markdown(path, report)
    if suffix == ".xml":
        return _rows_from_ipxact(path, report)
    raise ValueError(f"unsupported register input: {path}")


def _row_to_reg(row: dict[str, Any], report: ParseReport) -> tuple[tuple[str, int], dict, dict]:
    width = _parse_int(row.get("width"), 32)
    name = _sv_id(row.get("name"), "REG")
    offset = _parse_int(row.get("offset"))
    source = row.get("_source", "row")
    reg_field_access = _normalize_field_access(
        row.get("access"), "RW", report, f"{source} {name}.access"
    )
    access = _access_to_map_right(reg_field_access)
    reset = _parse_int(row.get("reset"), 0)
    field_name = _sv_id(row.get("field_name"), "data")
    field_bits = _normalize_bits(row.get("field_bits"), width)
    field_access = _normalize_field_access(
        row.get("field_access"),
        reg_field_access,
        report,
        f"{source} {name}.{field_name}.access",
    )
    # G8: field-level reset wins only when explicitly given. When the row
    # omits it, derive the field's reset by slicing the register-level reset
    # for this field's bit range — defaulting to 0 would turn an absent
    # field reset into an explicit zero and trigger false reset mismatches.
    raw_field_reset = row.get("field_reset")
    if raw_field_reset in (None, ""):
        f_lsb, f_width = _bit_lsb_width(field_bits)
        field_reset = (reset >> f_lsb) & ((1 << f_width) - 1)
    else:
        field_reset = _parse_int(raw_field_reset, 0)
    reg = {
        "name": name,
        "offset": offset,
        "width": width,
        "access": access,
        "reset": reset,
        "fields": [],
    }
    for opt in ("aliased_by", "aliased_by_value", "array_of", "stride", "effect"):
        if row.get(opt) not in (None, ""):
            reg[opt] = _parse_int(row[opt]) if opt in {"aliased_by_value", "array_of", "stride"} else row[opt]
    if row.get("effect") not in (None, ""):
        report.side_effects.append(f"{name}: effect={row['effect']} from {row.get('_source', 'input')}")
    field = {
        "name": field_name,
        "bits": field_bits,
        "access": field_access,
        "reset": field_reset,
    }
    if row.get("desc"):
        field["desc"] = str(row["desc"])
    return (name, offset), reg, field


def _collect_regs(rows: list[dict[str, Any]], report: ParseReport) -> list[dict]:
    regs: OrderedDict[tuple[str, int], dict] = OrderedDict()
    for row in rows:
        try:
            key, reg, field = _row_to_reg(row, report)
        except Exception as exc:
            report.warnings.append(f"{row.get('_source', 'row')}: skipped row: {exc}")
            continue
        if key not in regs:
            regs[key] = reg
        else:
            existing = regs[key]
            for attr in ("width", "access", "reset"):
                if existing[attr] != reg[attr]:
                    report.conflicts.append(
                        f"{reg['name']}@0x{reg['offset']:X}: {attr} {existing[attr]!r} vs {reg[attr]!r}"
                    )
            for opt in ("aliased_by", "aliased_by_value", "array_of", "stride", "effect"):
                if opt in reg and opt not in existing:
                    existing[opt] = reg[opt]
        regs[key]["fields"].append(field)
    for reg in regs.values():
        field_summary = _fields_to_map_right(reg["fields"])
        if reg["access"] != field_summary:
            report.access_normalization.append(
                f"{reg['name']}@0x{reg['offset']:X}: register access summary "
                f"{reg['access']} -> {field_summary} from fields"
            )
            reg["access"] = field_summary
    return list(regs.values())


def _same_array_shape(a: dict, b: dict) -> bool:
    return (
        a["width"] == b["width"]
        and a["access"] == b["access"]
        and a["reset"] == b["reset"]
        and len(a["fields"]) == len(b["fields"])
        and [
            (f["name"], f["bits"], f["access"], f["reset"]) for f in a["fields"]
        ] == [
            (f["name"], f["bits"], f["access"], f["reset"]) for f in b["fields"]
        ]
    )


def _fold_arrays(regs: list[dict], report: ParseReport) -> list[dict]:
    by_base: dict[str, list[tuple[int, int, dict]]] = {}
    for idx, reg in enumerate(regs):
        m = re.match(r"^(.+?)(\d+)$", reg["name"])
        if m:
            by_base.setdefault(m.group(1), []).append((int(m.group(2)), idx, reg))

    consumed: set[int] = set()
    folded_by_first: dict[int, dict] = {}
    for base, items in by_base.items():
        items = sorted(items)
        if len(items) < 2 or items[0][0] != 0:
            continue
        stride = items[1][2]["offset"] - items[0][2]["offset"]
        if stride <= 0:
            continue
        first = items[0][2]
        ok = True
        for pos, (_, _, reg) in enumerate(items):
            if reg["offset"] != first["offset"] + pos * stride or not _same_array_shape(first, reg):
                ok = False
                break
        if not ok:
            report.array_folding.append(f"{base}: not folded; non-contiguous or heterogeneous shape")
            continue
        folded = {k: v for k, v in first.items() if k != "fields"}
        folded["name"] = base
        folded["array_of"] = len(items)
        folded["stride"] = stride
        folded["fields"] = first["fields"]
        folded_by_first[items[0][1]] = folded
        consumed.update(idx for _, idx, _ in items)
        report.array_folding.append(f"{base}: folded {len(items)} registers stride={stride}")

    out: list[dict] = []
    for idx, reg in enumerate(regs):
        if idx in folded_by_first:
            out.append(folded_by_first[idx])
        elif idx not in consumed:
            out.append(reg)
    return out


def _validate_regs(regs: list[dict], report: ParseReport, paddr_width: int | None) -> None:
    by_offset: dict[int, list[dict]] = {}
    for reg in regs:
        by_offset.setdefault(reg["offset"], []).append(reg)
        field_reset = 0
        used_bits: set[int] = set()
        for field in reg["fields"]:
            lsb, width = _bit_lsb_width(field["bits"])
            for bit in range(lsb, lsb + width):
                if bit in used_bits:
                    report.warnings.append(f"{reg['name']}: overlapping field bit {bit}")
                used_bits.add(bit)
            field_reset |= _parse_int(field.get("reset", 0), 0) << lsb
        mask = (1 << reg["width"]) - 1
        if (field_reset & mask) != (reg["reset"] & mask):
            report.reset_mismatches.append(
                f"{reg['name']}: reg reset 0x{reg['reset']:X} != field-derived 0x{field_reset & mask:X}; field values win for RAL"
            )
        if paddr_width is not None and reg["offset"] >= (1 << paddr_width):
            report.warnings.append(f"{reg['name']}: offset 0x{reg['offset']:X} exceeds paddr width {paddr_width}")

    for offset, same in by_offset.items():
        if len(same) <= 1:
            continue
        accesses = sorted(r["access"] for r in same)
        if len(same) == 2 and accesses == ["RO", "WO"]:
            report.alias_decisions.append(f"0x{offset:X}: disjoint RO/WO pair {[r['name'] for r in same]}")
        elif any(r.get("aliased_by") for r in same):
            report.alias_decisions.append(f"0x{offset:X}: banked alias {[r['name'] for r in same]}")
        else:
            report.warnings.append(f"0x{offset:X}: repeated offset without aliased_by {[r['name'] for r in same]}")


def _write_registers_yaml(regs: list[dict], out_yaml: Path) -> None:
    out_yaml.parent.mkdir(parents=True, exist_ok=True)
    out_yaml.write_text(yaml.safe_dump({"registers": regs}, sort_keys=False))


def normalize_registers(
    inputs: list[Path],
    out_yaml: Path,
    report_path: Path | None = None,
    paddr_width: int | None = None,
) -> list[dict]:
    report = ParseReport()
    rows: list[dict[str, Any]] = []
    for path in inputs:
        rows.extend(_rows_for_input(path, report))
    if not rows:
        raise SystemExit("FATAL: no register rows parsed")
    regs = _collect_regs(rows, report)
    regs = _fold_arrays(regs, report)
    _validate_regs(regs, report, paddr_width)
    _write_registers_yaml(regs, out_yaml)
    if report_path is not None:
        report.write(report_path)
    return regs


def parse_xlsx_to_yaml(xlsx: Path, out_yaml: Path, report_path: Path | None = None) -> None:
    """Compatibility wrapper used by run_evals.py."""
    normalize_registers([xlsx], out_yaml, report_path)


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--input", action="append", type=Path, help="xlsx/csv/md/xml register source")
    ap.add_argument("--xlsx", type=Path, help="backward-compatible alias for --input")
    ap.add_argument("--out", required=True, type=Path)
    ap.add_argument("--parse-report", type=Path)
    ap.add_argument("--paddr-width", type=int)
    args = ap.parse_args()
    inputs = list(args.input or [])
    if args.xlsx:
        inputs.append(args.xlsx)
    if not inputs:
        ap.error("at least one --input or --xlsx is required")
    normalize_registers(inputs, args.out, args.parse_report, args.paddr_width)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
